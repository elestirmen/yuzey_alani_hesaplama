from __future__ import annotations

import json
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
import pandas as pd
import rasterio
from rasterio.crs import CRS
from rasterio.transform import from_origin

import main
from main import RunConfig


def _write_dem_geotiff(path: Path, z: np.ndarray, *, dx: float, dy: float, crs: CRS) -> None:
    rows, cols = z.shape
    profile = {
        "driver": "GTiff",
        "height": rows,
        "width": cols,
        "count": 1,
        "dtype": "float32",
        "crs": crs,
        "transform": from_origin(0.0, float(rows) * float(dy), float(dx), float(dy)),
        "nodata": None,
        "compress": "deflate",
        "predictor": 2,
        "tiled": False,
    }
    with rasterio.open(path, "w", **profile) as dst:
        dst.write(z.astype(np.float32, copy=False), 1)


def _demo_dem(rows: int, cols: int, *, dx: float, dy: float) -> np.ndarray:
    xs = (np.arange(cols, dtype=np.float64) + 0.5) * float(dx)
    ys = (np.arange(rows, dtype=np.float64) + 0.5) * float(dy)
    x_grid, y_grid = np.meshgrid(xs, ys)
    plane = 0.08 * x_grid - 0.03 * y_grid + 10.0
    waves = 0.6 * np.sin(x_grid / 3.0) * np.cos(y_grid / 4.0)
    return (plane + waves).astype(np.float64, copy=False)


def _read_results_sheet(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(workbook_path, sheet_name=sheet_name)


def _write_roi_geojson(path: Path) -> None:
    payload = {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "properties": {"id": "roi_a"},
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [[[2.0, 2.0], [12.0, 2.0], [12.0, 12.0], [2.0, 12.0], [2.0, 2.0]]],
                },
            }
        ],
    }
    path.write_text(json.dumps(payload), encoding="utf-8")


def test_runconfig_to_argv_includes_workers(tmp_path: Path) -> None:
    dem_path = tmp_path / "dem.tif"
    _write_dem_geotiff(
        dem_path,
        np.zeros((8, 8), dtype=np.float64),
        dx=1.0,
        dy=1.0,
        crs=CRS.from_epsg(3857),
    )

    cfg = RunConfig(
        dem=str(dem_path),
        outdir=str(tmp_path / "out"),
        gsd=[1.0, 2.0],
        methods=["gradient_multiplier"],
        plots=False,
        workers=3,
    )
    argv = cfg.to_argv()

    assert "--workers" in argv
    assert argv[argv.index("--workers") + 1] == "3"


def test_runconfig_to_argv_supports_native_gsd(tmp_path: Path) -> None:
    dem_path = tmp_path / "dem_native.tif"
    _write_dem_geotiff(
        dem_path,
        np.zeros((8, 8), dtype=np.float64),
        dx=1.0,
        dy=1.0,
        crs=CRS.from_epsg(3857),
    )

    cfg = RunConfig(
        dem=str(dem_path),
        outdir=str(tmp_path / "out_native"),
        gsd=["native", 2.0],
        methods=["gradient_multiplier"],
        plots=False,
    )
    argv = cfg.to_argv()

    assert "--gsd" in argv
    assert argv[argv.index("--gsd") + 1 : argv.index("--gsd") + 3] == ["native", "2"]


def test_cli_parallel_workers_matches_serial(tmp_path: Path) -> None:
    from surface_area.cli import main as cli_main

    dem_path = tmp_path / "dem.tif"
    out_serial = tmp_path / "out_serial"
    out_parallel = tmp_path / "out_parallel"
    z = _demo_dem(40, 40, dx=1.0, dy=1.0)
    _write_dem_geotiff(dem_path, z, dx=1.0, dy=1.0, crs=CRS.from_epsg(3857))

    base_args = [
        "run",
        "--dem",
        str(dem_path),
        "--gsd",
        "1",
        "2",
        "--methods",
        "gradient_multiplier",
        "tin_2tri_cell",
    ]

    rc_serial = cli_main([*base_args, "--outdir", str(out_serial), "--workers", "1"])
    rc_parallel = cli_main([*base_args, "--outdir", str(out_parallel), "--workers", "2"])

    assert rc_serial == 0
    assert rc_parallel == 0

    cols = ["gsd_m", "method", "dx", "dy", "A2D", "A3D", "ratio", "valid_cells", "note"]
    df_serial = _read_results_sheet(out_serial / "results.xlsx", "results_long")[cols].sort_values(
        ["gsd_m", "method"]
    ).reset_index(
        drop=True
    )
    df_parallel = _read_results_sheet(out_parallel / "results.xlsx", "results_long")[cols].sort_values(
        ["gsd_m", "method"]
    ).reset_index(
        drop=True
    )

    assert list(df_serial["method"]) == list(df_parallel["method"])
    assert list(df_serial["valid_cells"]) == list(df_parallel["valid_cells"])
    assert list(df_serial["note"]) == list(df_parallel["note"])
    for col in ["gsd_m", "dx", "dy", "A2D", "A3D", "ratio"]:
        assert np.allclose(df_serial[col].to_numpy(), df_parallel[col].to_numpy(), rtol=0.0, atol=1e-9)

    info_parallel = json.loads((out_parallel / "run_info.json").read_text(encoding="utf-8"))
    assert info_parallel["params"]["workers"] == 2


def test_cli_native_gsd_uses_source_grid_without_resampling(tmp_path: Path) -> None:
    from surface_area.cli import main as cli_main

    dem_path = tmp_path / "dem_native_grid.tif"
    outdir = tmp_path / "out_native_grid"
    z = _demo_dem(30, 24, dx=2.0, dy=3.0)
    _write_dem_geotiff(dem_path, z, dx=2.0, dy=3.0, crs=CRS.from_epsg(3857))

    rc = cli_main(
        [
            "run",
            "--dem",
            str(dem_path),
            "--outdir",
            str(outdir),
            "--gsd",
            "native",
            "--methods",
            "gradient_multiplier",
        ]
    )

    assert rc == 0

    df = _read_results_sheet(outdir / "results.xlsx", "results_long")
    assert len(df) == 1
    row = df.iloc[0]
    assert np.isclose(float(row["dx"]), 2.0, rtol=0.0, atol=1e-12)
    assert np.isclose(float(row["dy"]), 3.0, rtol=0.0, atol=1e-12)
    assert "resampling=native" in str(row["note"])

    info_payload = json.loads((outdir / "run_info.json").read_text(encoding="utf-8"))
    assert info_payload["params"]["gsd_specs"] == ["native"]


def test_main_run_includes_synthetic_ground_truth_columns_when_reference_sidecar_exists(
    tmp_path: Path,
    monkeypatch,
) -> None:
    import generate_synthetic_tif

    dem_path = tmp_path / "analytic_dem.tif"
    outdir = tmp_path / "out_with_gt"

    rc_gen = generate_synthetic_tif.main(
        [
            "--out",
            str(dem_path),
            "--target",
            "analytic_gaussian_hill",
            "--dx",
            "1",
            "--extent_width",
            "24",
            "--extent_height",
            "18",
            "--seed",
            "5",
            "--eval_gsd",
            "2",
            "--quiet",
        ]
    )
    assert rc_gen == 0

    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "run",
            "--dem",
            str(dem_path),
            "--outdir",
            str(outdir),
            "--gsd",
            "native",
            "2",
            "--methods",
            "gradient_multiplier",
        ],
    )

    rc = main.main()
    assert rc == 0

    df = _read_results_sheet(outdir / "results.xlsx", "results_long").sort_values(["gsd_m", "method"]).reset_index(
        drop=True
    )
    assert "continuous_gt_A3D" in df.columns
    assert "A3D_continuous_gt_rel_err" in df.columns
    assert "synthetic_native_ref_A3D" in df.columns
    assert "A3D_synthetic_native_ref_rel_err" in df.columns
    assert {"continuous_ground_truth", "native_grid_reference"}.issubset(set(df["method"]))

    computed_rows = df[~df["method"].isin(["continuous_ground_truth", "native_grid_reference"])].reset_index(drop=True)
    assert computed_rows["synthetic_native_ref_A3D"].notna().all()

    xls = pd.ExcelFile(outdir / "results.xlsx")
    assert "reference_summary" in set(xls.sheet_names)
    df_ref = _read_results_sheet(outdir / "results.xlsx", "reference_summary")
    assert set(df_ref["reference_kind"]) == {"continuous_ground_truth", "native_grid_reference"}
    assert len(df_ref[df_ref["reference_kind"] == "continuous_ground_truth"]) == 1
    native_refs = df_ref[df_ref["reference_kind"] == "native_grid_reference"].sort_values("gsd_m").reset_index(drop=True)
    assert np.allclose(native_refs["gsd_m"].to_numpy(), [1.0, 2.0], rtol=0.0, atol=1e-12)
    assert native_refs["reference_A3D"].notna().all()

    sidecar = json.loads(dem_path.with_suffix(".reference.json").read_text(encoding="utf-8"))
    gt_a3d = float(sidecar["continuous_ground_truth"]["surface_area_m2"])
    assert np.allclose(computed_rows["continuous_gt_A3D"].to_numpy(), gt_a3d, rtol=0.0, atol=1e-9)
    continuous_rows = df[df["method"] == "continuous_ground_truth"].sort_values("gsd_m").reset_index(drop=True)
    assert np.allclose(continuous_rows["A3D"].to_numpy(), [gt_a3d, gt_a3d], rtol=0.0, atol=1e-9)
    continuous_summary = df_ref[df_ref["reference_kind"] == "continuous_ground_truth"].iloc[0]
    assert np.isclose(float(continuous_summary["reference_A3D"]), gt_a3d, rtol=0.0, atol=1e-9)

    run_info = json.loads((outdir / "run_info.json").read_text(encoding="utf-8"))
    assert run_info["synthetic_reference"]["has_continuous_ground_truth"] is True
    assert run_info["synthetic_reference"]["terrain_family"] == "analytic"


def test_cli_wide_results_only_include_calculated_method_areas(tmp_path: Path) -> None:
    from surface_area.cli import main as cli_main

    dem_path = tmp_path / "dem_wide.tif"
    outdir = tmp_path / "out_wide"
    z = _demo_dem(30, 30, dx=1.0, dy=1.0)
    _write_dem_geotiff(dem_path, z, dx=1.0, dy=1.0, crs=CRS.from_epsg(3857))

    rc = cli_main(
        [
            "run",
            "--dem",
            str(dem_path),
            "--outdir",
            str(outdir),
            "--gsd",
            "1",
            "2",
            "--methods",
            "gradient_multiplier",
            "tin_2tri_cell",
        ]
    )

    assert rc == 0

    workbook_path = outdir / "results.xlsx"
    assert workbook_path.exists()

    xls = pd.ExcelFile(workbook_path)
    assert set(xls.sheet_names) == {"results_long", "results_wide", "run_info", "grafikler"}

    df_wide = _read_results_sheet(workbook_path, "results_wide").sort_values("gsd_m").reset_index(drop=True)
    assert list(df_wide.columns) == ["gsd_m", "gradient_multiplier_A3D", "tin_2tri_cell_A3D"]

    df_long = _read_results_sheet(workbook_path, "results_long")
    expected = (
        df_long.pivot(index="gsd_m", columns="method", values="A3D")
        .rename(columns=lambda method: f"{method}_A3D")
        .reset_index()
        .sort_values("gsd_m")
        .reset_index(drop=True)
    )

    assert np.allclose(df_wide["gsd_m"].to_numpy(), expected["gsd_m"].to_numpy(), rtol=0.0, atol=1e-12)
    assert np.allclose(
        df_wide["gradient_multiplier_A3D"].to_numpy(),
        expected["gradient_multiplier_A3D"].to_numpy(),
        rtol=0.0,
        atol=1e-9,
    )
    assert np.allclose(
        df_wide["tin_2tri_cell_A3D"].to_numpy(),
        expected["tin_2tri_cell_A3D"].to_numpy(),
        rtol=0.0,
        atol=1e-9,
    )

    workbook = load_workbook(workbook_path)
    charts_ws = workbook["grafikler"]
    assert charts_ws["A1"].value == "Bu sayfadaki grafikler Excel icinde duzenlenebilir; kaynak tablolar asagidadir."
    assert charts_ws["A68"].value == "A3D vs GSD - kaynak veri"
    assert charts_ws.max_row >= 70
    assert len(charts_ws._charts) >= 3


def test_cli_multiscale_only_results_long_column_order_is_stable(tmp_path: Path) -> None:
    from surface_area.cli import main as cli_main

    dem_path = tmp_path / "dem_multiscale.tif"
    outdir = tmp_path / "out_multiscale"
    z = _demo_dem(24, 24, dx=1.0, dy=1.0)
    _write_dem_geotiff(dem_path, z, dx=1.0, dy=1.0, crs=CRS.from_epsg(3857))

    rc = cli_main(
        [
            "run",
            "--dem",
            str(dem_path),
            "--outdir",
            str(outdir),
            "--gsd",
            "1",
            "--methods",
            "multiscale_decomposed_area",
        ]
    )

    assert rc == 0

    df_long = _read_results_sheet(outdir / "results.xlsx", "results_long")
    assert list(df_long.columns) == [
        "gsd_m",
        "dx",
        "dy",
        "method",
        "A2D",
        "A3D",
        "ratio",
        "valid_cells",
        "runtime_sec",
        "resample_runtime_sec",
        "note",
        "a_topo",
        "a_micro",
        "a_total",
        "micro_ratio",
        "sigma_m",
    ]
    assert "note" in df_long.columns
    assert df_long.columns.get_loc("note") < df_long.columns.get_loc("a_topo")


def test_cli_roi_results_sheet_column_order_is_consistent(tmp_path: Path) -> None:
    from surface_area.cli import main as cli_main

    dem_path = tmp_path / "dem_roi.tif"
    roi_path = tmp_path / "roi.geojson"
    outdir = tmp_path / "out_roi"
    z = _demo_dem(30, 30, dx=1.0, dy=1.0)
    _write_dem_geotiff(dem_path, z, dx=1.0, dy=1.0, crs=CRS.from_epsg(3857))
    _write_roi_geojson(roi_path)

    rc = cli_main(
        [
            "run",
            "--dem",
            str(dem_path),
            "--outdir",
            str(outdir),
            "--gsd",
            "1",
            "--methods",
            "adaptive_bilinear_patch_integral",
            "sector_adaptive_jenness_integral",
            "--roi",
            str(roi_path),
            "--roi_id_field",
            "id",
        ]
    )

    assert rc == 0

    df_roi = _read_results_sheet(outdir / "results.xlsx", "results_roi_long")
    assert list(df_roi.columns) == [
        "gsd_m",
        "dx",
        "dy",
        "roi_id",
        "method",
        "A2D",
        "A3D",
        "ratio",
        "valid_cells",
        "runtime_sec",
        "resample_runtime_sec",
        "note",
        "adaptive_avg_level",
        "adaptive_max_level_used",
        "adaptive_refined_cell_fraction",
        "adaptive_total_subcells_evaluated",
        "sector_jenness_avg_level",
        "sector_jenness_max_level_used",
        "sector_jenness_refined_fraction",
    ]


def test_compute_methods_on_raster_workers_matches_serial(tmp_path: Path) -> None:
    from surface_area.methods import compute_methods_on_raster

    dem_path = tmp_path / "dem_methods.tif"
    z = _demo_dem(36, 36, dx=1.0, dy=1.0)
    _write_dem_geotiff(dem_path, z, dx=1.0, dy=1.0, crs=CRS.from_epsg(3857))

    serial = compute_methods_on_raster(
        str(dem_path),
        nodata=None,
        methods=["gradient_multiplier", "tin_2tri_cell", "adaptive_bilinear_patch_integral"],
        jenness_weight=0.25,
        slope_method="horn",
        integral_N=5,
        workers=1,
    )
    parallel = compute_methods_on_raster(
        str(dem_path),
        nodata=None,
        methods=["gradient_multiplier", "tin_2tri_cell", "adaptive_bilinear_patch_integral"],
        jenness_weight=0.25,
        slope_method="horn",
        integral_N=5,
        workers=2,
    )

    assert serial.keys() == parallel.keys()
    for key in serial:
        s = serial[key]
        p = parallel[key]
        assert s.valid_cells == p.valid_cells
        assert np.isclose(s.a3d, p.a3d, rtol=0.0, atol=1e-9)
        assert s.adaptive_max_level_used == p.adaptive_max_level_used
        assert s.adaptive_total_subcells_evaluated == p.adaptive_total_subcells_evaluated
